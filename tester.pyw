#!/usr/bin/env python3
"""Resistance Measuring software"""

import csv
import json

from datetime import datetime
from functools import partial
from glob import glob
from os import path
from queue import Queue
from random import randint
from threading import Thread
from time import sleep
from tkinter import Tk, Menu, Entry, LabelFrame, Label, \
        StringVar, IntVar, messagebox as mBox, filedialog
from tkinter.ttk import Style, Combobox, Button, Checkbutton

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from serial import Serial, SerialException
from serial.tools.list_ports import comports

VERSION             = '1.2'
TITLE               = 'Resistance Measuring V{}'.format(VERSION)

LOOP_TIME           = 150     # milliseconds
SERIAL_READ_TIMEOUT = 2.5


def current_path(filename):
    """Get current path of app

    Args:
        filename (string): xxxx.xxx

    Returns:
        string: full path
    """

    return  path.join(
            path.dirname(path.realpath(__file__)),
            filename)


class Config:
    """Save a configuration"""

    _path = current_path('config.json')

    comport = None
    test_wo_sensor = False
    lot_no = ''
    upper_bound = 400.0
    lower_bound = 300.0

    def __init__(self):

        if not path.exists(self._path):
            return

        with open(self._path) as readfile:
            obj = json.load(readfile)

            if 'comport' in obj:
                self.comport = obj['comport']
            if 'test_wo_sensor' in obj:
                self.test_wo_sensor = obj['test_wo_sensor']
            if 'lot_no' in obj:
                self.lot_no = obj['lot_no']
            if 'upper_bound' in obj:
                self.upper_bound = obj['upper_bound']
            if 'lower_bound' in obj:
                self.lower_bound = obj['lower_bound']

    def save(self):
        """Save to json file"""

        with open(self._path, 'w') as outfile:
            out = {
                'comport': self.comport,
                'test_wo_sensor': self.test_wo_sensor,
                'lot_no': self.lot_no,
                'upper_bound': self.upper_bound,
                'lower_bound': self.lower_bound,
            }
            json.dump(out, outfile, indent=2)


class Recorder:
    """Record resistance value into csv log file
    """

    def record(self, lot, val, result='', cable_no=None):
        """Record value

        Args:
            lot (string): lot number
            val (float): resistance value
            result (string): resistance value
            cable_no (number): cable number

        Returns:
            string: filename
        """

        time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        _path = current_path('{}.csv'.format(lot))

        with open(_path, 'a+') as outfile:
            if cable_no:
                msg = '{},{},{},{},{}\n'.format(time, lot, cable_no, val, result)
            else:
                msg = '{},{},,{},{}\n'.format(time, lot, val, result)
            print('recording: {}'.format(msg), end='')

            outfile.write(msg)


    def get_last_cable_number(self, lot):
        """Get last cable number from a lot

        Args:
            lot (string): lot name

        Returns:
            number: last number
        """

        _path = current_path('{}.csv'.format(lot))

        number = 0
        if not path.exists(_path):
            return number

        with open(_path) as infile:
            lines = infile.readlines()
            cols = lines[-1].split(',')
            if cols[2].isdigit():
                # cable number
                number = int(cols[2])

        return number


class MainApp(Tk):
    """Main Window"""

    cfg = Config()
    recorder = Recorder()
    queue = Queue()
    serial = None
    last_month = ''

    def __init__(self, win):
        """Big frame"""

        self.master = win

        menu_bar = Menu(win)
        win.config(menu=menu_bar)

        file_menu = Menu()
        file_menu.add_separator()
        file_menu.add_command(label='Exit', command=self.quit)
        menu_bar.add_cascade(label='File', menu=file_menu)

        self.resistance_label = Label(win, text='XXX')
        self.resistance_label.config(background='green', foreground='black')
        self.resistance_label.grid(
                column=0, row=0,
                sticky='nsew',
                padx=12, pady=10)
        win.rowconfigure(0, weight=1)
        win.columnconfigure(0, weight=1)

        s = Style()
        s.configure('my.TButton', font=(18))
        self.check_button = Button(win, text='\nCheck\n',
                command=self.check,
                style='my.TButton',
                )
        self.check_button.grid(column=0, row=1)
        self.check_button.config(width=20)
        self.check_button.focus()

        setting_frame = LabelFrame(win, text='Setting')
        setting_frame.grid(row=2, column=0, pady=10)

        self.selected_port_var = StringVar()
        self.port_combobox = Combobox(setting_frame, state='readonly',
                textvariable=self.selected_port_var)
        self.port_combobox.grid(row=0, column=0, padx=8)
        self._update_portlist()

        Button(setting_frame,
                text='Refresh',
                command=self._update_portlist,
                ).grid(row=1, column=0)

        Label(setting_frame, text='Lot no.').grid(row=0, column=1)
        self.lot_var = StringVar()
        self.lot_entry = Entry(setting_frame, width=20,
                state='readonly', textvariable=self.lot_var,)
        self.lot_entry.grid(row=0, column=2, padx=8, pady=4)

        self.lot_lock_var = IntVar()
        Checkbutton(setting_frame, text="Lock",
                variable=self.lot_lock_var,
                command=self.on_lock_changed,
                ).grid(row=0, column=3)

        Label(setting_frame, text='Cable no.').grid(row=1, column=1)
        self.cable_var = StringVar()
        cable_entry = Entry(setting_frame, width=20, textvariable=self.cable_var, justify='center', state='readonly')
        cable_entry.grid(row=1, column=2, pady=4)
        self.export_button = Button(setting_frame,
                text='Export',
                command=self._export_data,
                )
        self.export_button.grid(row=1, column=3, padx=8)

        #
        # lower / upper tolerance
        #
        Label(setting_frame, text='Upper bound').grid(row=2, column=1)
        self.upper_var = StringVar()
        self.upper_entry = Entry(setting_frame, width=20,
                textvariable=self.upper_var, justify='center', state='readonly')
        self.upper_entry.grid(row=2, column=2)

        Label(setting_frame, text='Lower bound').grid(row=3, column=1)
        self.lower_var = StringVar()
        self.lower_entry = Entry(setting_frame, width=20,
                textvariable=self.lower_var, justify='center', state='readonly')
        self.lower_entry.grid(row=3, column=2)


        #
        # init cable info
        #
        self.upper_var.set(self.cfg.upper_bound)
        self.lower_var.set(self.cfg.lower_bound)
        self.lot_var.set(self.cfg.lot_no)
        self.lot_lock_var.set(1)
        self.cable_var.set(self.recorder.get_last_cable_number(self.cfg.lot_no))

        win.bind("<Configure>", self.on_resize)


    def on_resize(self, event):
        """Window changed size"""

        #print('on_resize: {} {}'.format(event.width, event.height))
        self.resistance_label.config(
                font=('times', int(self.master.winfo_width() / 5), 'bold'))


    def on_lock_changed(self):
        """Lot number lock has been changed"""

        if self.lot_lock_var.get():
            self.check_button.config(state='active')
            self.export_button.config(state='active')
            self.lot_entry.config(state='readonly')
            self.upper_entry.config(state='readonly')
            self.lower_entry.config(state='readonly')
            cable_no = self.recorder.get_last_cable_number(self.lot_var.get())
            self.cable_var.set(str(cable_no))

            self.cfg.upper_bound = float(self.upper_var.get())
            self.cfg.lower_bound = float(self.lower_var.get())
        else:
            self.check_button.config(state='disabled')
            self.export_button.config(state='disabled')
            self.lot_entry.config(state='normal')
            self.upper_entry.config(state='normal')
            self.lower_entry.config(state='normal')


    def quit(self):
        """Exit app"""

        self.master.quit()
        self.master.destroy()
        exit()


    def check(self):
        """Check button pressed"""

        #
        # Manage ui
        #
        self.resistance_label.config(bg='gray', text='?')
        self.export_button.config(state='disabled')
        self.check_button.config(state='disabled')
        self.check_button.focus()

        self.queue.queue.clear()
        self._thread = Thread(target=self._get_resistance,
                name='resistance', daemon=True)
        self._thread.start()
        self.master.after(LOOP_TIME, self._listen_for_resistance)

        ##
        ## Save selected comport
        ##
        #selected_port = self.selected_port.get()
        #if selected_port:
            #self.cfg.comport = selected_port
            #self.cfg.save()


    def _get_resistance(self):
        """Get resistance from serial port"""

        if self.cfg.test_wo_sensor:
            sleep(1)

            self.cfg.upper_bound = float(self.upper_var.get())
            upper = int(self.cfg.upper_bound * 125)
            lower = int(self.cfg.lower_bound * 75)
            val = randint(lower, upper) / 100

        else:
            try:
                port = self.selected_port_var.get()
                if not self.serial:
                    self.serial =  Serial(port, 9600, timeout=SERIAL_READ_TIMEOUT)
                    self.serial.readline()
                elif port != self.serial.name:
                    self.serial.close()
                    self.serial =  Serial(port, 9600, timeout=SERIAL_READ_TIMEOUT)
                    self.serial.readline()

                self.serial.flush()
                self.serial.write(b'M\n')
                data = self.serial.readline()
                if len(data) == 0:
                    data = None

                msg = data.decode('ascii').strip()
                #print('{}: {} -> {}'.format(self.serial.name, data, msg))

                val = float(msg)

                if port != self.cfg.comport:
                    self.cfg.comport = port
                    self.cfg.save()

            except SerialException as e:
                print(e)
                self.serial = None
                val = -1
            except AttributeError as e:
                print(e)
                val = -2
            except ValueError as e:
                print(e)
                val = -3
                if port != self.cfg.comport:
                    self.cfg.comport = port
                    self.cfg.save()
            except Exception as e:
                print('{}: {}'.format(type(e), e))
                self.serial = None
                val = -10

        self.queue.put(val)


    def _listen_for_resistance(self):
        """Get returned value from sensor reading"""

        if self._thread.is_alive():
            self.master.after(LOOP_TIME, self._listen_for_resistance)
            return

        val = self.queue.get()
        lot_no, cable_no = self._get_cable_info()

        result = 'Pass'
        if val >= self.cfg.lower_bound and val <= self.cfg.upper_bound:
            self.resistance_label.config(bg='green', text='{}'.format(val))
        else:
            result = 'Fail'
            self.resistance_label.config(bg='red', text='{}'.format(val))

        if val > 0:
            self.recorder.record(lot_no, val, result, cable_no)

            self.cfg.lot_no = lot_no
            self.cfg.save()
            self.cable_var.set(str(cable_no))

        else:
            self._error_dialog(val)

        self.export_button.config(state='active')
        self.check_button.config(state='active')


    def _get_cable_info(self):
        """Get cable lot, number

        Returns:
            string: cable lot
            number: cable number
        """

        lot_no = self.lot_var.get().strip()
        if len(lot_no) < 1:
            lot_no = 'xxx'
            self.lot_var.set(lot_no)

        try:
            cable_no = int(self.cable_var.get()) + 1
        except Exception as e:
            print(e)
            cable_no = 1

        return lot_no, cable_no


    def _update_portlist(self):
        """Re-assign port list"""

        ports = self._portlist()
        self.port_combobox['values'] = ports
        if self.cfg.comport and self.cfg.comport in ports:
            self.selected_port_var.set(self.cfg.comport)
        else:
            self.selected_port_var.set('')


    def _portlist(self):
        """Scan avaiable ports"""

        ports = []

        ##
        ## dummy ports
        ##
        #val = randint(1, 10)
        #for i in range(val):
            #ports.append('com{}'.format(i+1))
        #return ports

        _ports = comports()

        if len(_ports) < 1:
            return ports

        return [ port.device for port in _ports ]


    def _export_data(self):
        """Export lot log as xlsx
        """

        folder = filedialog.askdirectory()
        filename = self.lot_var.get()
        if not folder:
            return

        wb = Workbook()
        ws = wb.active
        headers = ('Time', 'Lot No.', 'Cable No.', 'Value', 'Result')
        no_cols = []
        for name in headers:
            if name in ('Cable No.', 'Value'):
                no_cols.append(headers.index(name))
        with open('{}.csv'.format(filename), 'r') as f:
            ws.append(headers)
            for row in csv.reader(f):
                for i in no_cols:
                    #
                    # Parse string to number for some columns
                    #
                    try:
                        row[i] = int(row[i])
                    except Exception as e:
                        row[i] = float(row[i])

                ws.append(row)

        fill_format = PatternFill(start_color="e0efd4", end_color="e0efd4", fill_type = "solid")
        self._adjust_column_width(ws)
        for i in range(len(headers)):
            ws['{}1'.format(chr(ord('A')+i))].fill = fill_format

        self._set_failed_row(ws)

        #
        # Freeze first row / col
        #
        c = ws['B2']
        ws.freeze_panes = c

        _path = path.join(folder, 'resistance-{}.xlsx'.format(filename))
        wb.save(_path)


    def _adjust_column_width(self, worksheet):
        """Manage columns of exported file

        Args:
            worksheet (object): current worksheet
        """

        adjusted_width = 0
        for col in worksheet.columns:
            max_length = 0
            col_idx = col[0].column # Get the column index
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
                cell.alignment = Alignment(horizontal='center')
            adjusted_width = (max_length + 2) * 1.2
            column_name = chr(ord('A') + col_idx - 1)
            worksheet.column_dimensions[column_name].width = adjusted_width


    def _set_failed_row(self, worksheet):
        """Make red on failed row

        Args:
            worksheet (Worksheet): active sheet
        """

        for row in worksheet.rows:
            if 'fail' in str(row[4].value).lower():
                for cell in row:
                    cell.font = Font(color = 'FFFF0000')

    def _error_dialog(self, err_no, title="เกิดข้อผิดพลาด"):
        """Show about box

        Args:
            err_no (number): error number
            title (string): title
        """

        print("_error_dialog {title}: {err_no}".format(
            title=title, err_no=err_no))
        if err_no == -1:
            title = 'No devices'
            message = u'ไม่เจออุปกรณ์'
        elif err_no == -2:
            title = 'No response'
            message = u'ไม่สามารถรับค่าได้'
        elif err_no == -3:
            title = 'Wrong format'
            message = u'ข้อมูลไม่ถูกต้อง'
        else:
            message = u'ไม่เจออุปกรณ์ ({})'.format(err_no)

        mBox.showerror(title, "ERROR: {} !".format(message))


if __name__ == '__main__':

    win = Tk()
    app = MainApp(win)

    win.title(TITLE)
    win.geometry('%dx%d+%d+%d' % (800, 550, 64, 32))
    # win.resizable(0, 0)

    win.mainloop()
